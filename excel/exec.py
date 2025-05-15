import openpyxl


def edit(path, num, RowOrCol, removestr):

    wb = openpyxl.load_workbook(path)

    for sheet in wb:

        if RowOrCol == "row":
            cols = sheet[num]
            now = 1
            for cells in cols:

                if str(cells.value) != str(removestr):

                    tmp = 1
                    for val in sheet[cells.column_letter]:
                        sheet.cell(row=tmp, column=now).value = val.value
                        tmp += 1
                    now += 1
            for i in range(now, len(cols) + 1):
                tmp = 1
                for dele in sheet[openpyxl.utils.get_column_letter(i)]:
                    sheet.cell(row=tmp, column=i).value = ""
                    tmp += 1

        else:

            rows = sheet[num]

            now = 1
            for cells in rows:

                if str(cells.value) != str(removestr):
                    tmp = 1
                    for val in sheet[cells.row]:
                        sheet.cell(row=now, column=tmp).value = val.value
                        tmp += 1
                    now += 1
            for i in range(now, len(rows) + 1):
                tmp = 1
                for dele in sheet[i]:
                    sheet.cell(row=i, column=tmp).value = ""
                    tmp += 1

    wb.save(path)
    wb.close()
